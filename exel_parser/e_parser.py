from openpyxl.cell.cell import Cell
import re
import csv
from typing import Dict, List, Union, Tuple, Generator
from openpyxl.worksheet.worksheet import Worksheet
import logging


CSV_PATH = 'result.csv'

with open(CSV_PATH, 'r') as csv_file:
    reader = csv.reader(csv_file)
    FIELDS = list(reader)[0]


logging.basicConfig(filename='bad_values.log', filemode='w',
                    format='%(asctime)s - %(message)s', level=logging.INFO)


class ExcelParser:
    """
    Класс для обработки данных из worksheet и их записи в csv файл.
    """
    csv_dict = {
        field: "" for field in FIELDS
    }

    row_dict = {}

    user_additional_info = ()
    rows_to_write = []

    def __init__(self, worksheet: Worksheet):
        self.worksheet = worksheet
        self.excel_fields_dict = self.set_fields()

    def set_fields(self) -> Dict[int, str]:
        col_values = (col[0].value.lower() for col in self.worksheet.iter_cols(1, self.worksheet.max_column))

        fields_dict = {
            column + 1: field_name for column, field_name in enumerate(col_values)
        }
        return fields_dict

    def generate_csv(self):
        for i in range(1, self.worksheet.max_row):
            col_values = (col[i].value for col in self.worksheet.iter_cols(1, self.worksheet.max_column))
            self.process_row(col_values)

        self.write_rows()

    def process_row(self, values_lst: Generator):
        for ind, value in enumerate(values_lst):
            if ind + 1 in self.excel_fields_dict.keys():
                self.row_dict[self.excel_fields_dict[ind + 1]] = str(value)

        self.process_row_dict()

    def process_row_dict(self):
        self.process_names()
        self.process_ssn()
        self.process_address()
        self.process_company()
        self.process_zip()
        self.process_tel()
        self.row_dict = {}
        self.csv_dict['user_additional_info'] = "|".join(self.user_additional_info)
        self.user_additional_info = ()
        self.rows_to_write.append(self.csv_dict)
        self.refresh_csv_dict()

    def refresh_csv_dict(self):
        self.csv_dict = {
            field: "" for field in FIELDS
        }

    @classmethod
    def correct_number(cls, number):
        digits = re.findall(r'\d', number)
        if len(digits) == 11:
            digits = digits[1:]
        return "1 ({}{}{}) {}{}{} {}{}{}{}".format(*digits)

    def process_tel(self):
        """
        Предполагая, что должно быть 11 цифр (или 10 если без начальной единицы),
        которые могут быть разделены знаками "-" или пробелами. Первые три цифры (со второй по четвертую,
        если есть единица) могут быть в кавычках.
        """
        tel = self.row_dict['mobile number']
        pattern = r"(?:1(?:-|\s))?(?:\(?\d{3}\)?(?:\s|-))(?:\d{3}(?:\s|-))\d{4}"
        tel_check = re.fullmatch(pattern, tel)
        if tel_check:
            if ('(' in tel and ')' in tel) or ('(' not in tel and ')' not in tel):

                self.csv_dict['tel'] = self.correct_number(tel)
                return
        logging.info(f"Bad phone number: {tel}")

    def write_rows(self):
        with open(CSV_PATH, 'a') as file:
            writer = csv.DictWriter(file, fieldnames=FIELDS)
            for data in self.rows_to_write:
                writer.writerow(data)

    def process_zip(self):
        user_zip = self.row_dict['zip']
        if re.fullmatch(
            r"\d{5}(?:-\d{4})?", user_zip
        ):
            self.csv_dict['zip'] = user_zip
        else:
            logging.info(f"Bad zip: {user_zip}")

    def process_company(self):
        position = self.row_dict['position']
        company = self.row_dict['company']
        check_position = re.fullmatch(
            r"(?:[a-z]+)(?:\s[a-z]+)*",
            position
        )
        check_company = re.fullmatch(
            r"(?:[A-Z][A-Za-z']+)(?<!['-])(?:(?:(?:\s|-|,\s)(?!['-])[A-Za-z'-]+)(?<!['-]))*",
            company
        )
        str_to_append = None
        if check_company and check_position:
            str_to_append = f"Работает в компании {company} в отделе {self.row_dict['department']} " \
                            f"в должности {position}"
        elif check_company and check_position is None:
            logging.info(f"Bad position: {position}")
            str_to_append = f"Работает в компании {company} в отделе {self.row_dict['department']}"
        elif check_company is None and check_position:
            logging.info(f"Bad company: {company}")
            str_to_append = f"Работает в должности {position}"
        else:
            logging.info(f"Bad company: {company}")
            logging.info(f"Bad position: {position}")
        if str_to_append is not None:
            self.user_additional_info += str_to_append,

    def process_address(self):
        """
        Предполагая следующие свойства адреса:
        - может или нет начинаться с Apt./Suite и номером квартиры;
        - название улицы может состоять из более одного слова, начинающиеся на заглавные буквы или цифру
        (например, 05th Avenue) и может содержать цифры и символы "-&'.";
        - название города может состоять из более одного слова, каждое из которых должно начинаться с
        заглавной, и содержать символы ".'";
        - код штата состоит из двух заглавных букв;
        - zip-код состоит из пяти цифр и может сопровождаться дополнительно дефисом и четырьмя
        цифрами
        """
        address = self.row_dict['address']
        address_pattern = r"(?:(?:Apt\.\s)|(?:Suite\s)\d+\s)?\d+(?:\s[A-Z\d][A-Za-z\d.'&-]+)+,(?:\s[A-Z][A-Za-z.']+)+,\s[A-Z]{2}\s\d{5}(?:-\d{4})?"
        if re.fullmatch(address_pattern, address):
            self.csv_dict['address'] = address
        else:
            logging.info(f"Bad address: {address}")

    def process_ssn(self):
        """
        Предполагая, что валидный ssn имеет формат AAA-GG-SSSS и:
        - первые три цифры не могут быть 000, 666 или в диапазоне 900-999;
        - четвертая и пятая цифры должны быть в диапазоне от 01 до 99;
        - последние четыре цифры должны быть в диапазоне 0001 до 9999.
        """
        ssn = self.row_dict['ssn']
        if re.fullmatch(
                r"(?!666|000|9\d{2})\d{3}-(?!00)\d{2}-(?!0000)\d{4}",
                ssn
        ):
            self.user_additional_info += f"Номер SSN: {ssn}",
        else:
            logging.info(f"Bad ssn: {ssn}")

    def process_names(self):
        first_name = self.row_dict['first name']
        last_name = self.row_dict['last name']
        if re.fullmatch(r"[A-Z][a-z]+", first_name) and re.fullmatch(r"[A-Z](?:[A-Za-z']+(?!'))", last_name):
            self.csv_dict['user_fullname'] = f"{first_name} {last_name}"
        else:
            logging.info(f"Bad names: {first_name} {last_name}")









